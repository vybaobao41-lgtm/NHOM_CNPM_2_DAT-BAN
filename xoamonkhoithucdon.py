import openpyxl

file_path = r"C:\Users\ADMIN\Downloads\THUCDON.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

def delete_dish_by_name():
    ten_mon_can_xoa = input("\nNhập TÊN MÓN cần xoá: ").strip()

    if not ten_mon_can_xoa:
        print("✖ Tên món không được để trống!")
        return

    dong_mon_can_xoa = None

    for row in range(2, ws.max_row + 1):
        ten_mon_trong_excel = ws.cell(row=row, column=2).value
        if ten_mon_trong_excel and ten_mon_trong_excel.strip().lower() == ten_mon_can_xoa.lower():
            dong_mon_can_xoa = row
            break

    if dong_mon_can_xoa is None:
        print("✖ Món không tồn tại trong thực đơn!")
        return

    xac_nhan = input(
        f"⚠ Bạn chắc chắn muốn xoá TOÀN BỘ thông tin của món '{ten_mon_can_xoa}'? (Y/N): "
    ).strip().lower()

    if xac_nhan != "y":
        print("ℹ Đã huỷ thao tác xoá.")
        return

    ws.delete_rows(dong_mon_can_xoa)
    wb.save(file_path)
    print("✔ Đã xoá món khỏi thực đơn (tên + danh mục + giá).")
